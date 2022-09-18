import shutil
import zipfile
from PyPDF2 import PdfReader
import csv
from openpyxl import load_workbook
import pytest
import os


zip_obj = zipfile.ZipFile("resources/test.zip", 'w')
zip_obj.write("resources/1.pdf")
zip_obj.write("resources/2.csv")
zip_obj.write("resources/3.xlsx")
zip_obj.close()


def test_pdf():
    with zipfile.ZipFile('resources/test.zip') as myzip:
        with myzip.open("resources/1.pdf") as pdf_data:
            pdf_data = PdfReader(pdf_data)
            page = pdf_data.pages[0]
            text = page.extract_text()
            assert "Пример" in text


def test_csv():
    with zipfile.ZipFile('resources/test.zip') as myzip:
        myzip.extract("resources/2.csv")
        with open("resources/2.csv", 'r') as csv_file:
            table = csv.reader(csv_file, delimiter=";")
            for line_no, line in enumerate(table, 1):
                if line_no == 2:
                    assert 'Иванова' in line[1]


def test_xlsx():
    with zipfile.ZipFile('resources/test.zip') as myzip:
        with myzip.open("resources/3.xlsx") as xlsx_data:
            workbook = load_workbook(xlsx_data)
            sheet = workbook.active
            sheet = sheet.cell(row=3, column=2).value
            assert "Сергеев" in sheet


















